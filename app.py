from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)

@app.route("/")
def index():
    images = []
    excel = load_workbook("gallery.xlsx")
    page = excel["Лист1"]
    for row in page:
        url = row[0].value
        title = row[2].value
        pics_list = [title, url] 
        images.append(pics_list) 

    return render_template('index.html', images=images)

@app.route("/add")
def add():
    return render_template("add.html")

@app.route("/add-pic", methods=["POST"])
def add_pic():
    description = request.form.get("description")
    pic = request.form.get["pic"]
    title = request.form.get("title")
    excel = load_workbook("gallery.xlsx")
    page = excel["Лист1"]
    page.append([pic, description, title])
    excel.save('gallery.xlsx')

    return render_template("picadded.html")

@app.route("/details/<number>")
def details(number):
    excel = load_workbook("gallery.xlsx")
    page = excel["Лист1"]
    lst = page[str(number)]
    return render_template("details.html", lst=lst)

